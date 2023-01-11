import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Union
from models.abstracts.file_operations import FileOperations

from models.constants.database_actions import DatabaseActions


class ExcelFile(FileOperations):
    def __init__(self, file_path):
        self.file_path = file_path
        self.headers = ["email", "age", "password", "score", "games played"]

    def get_all_users(self) -> List:
        excel_file = self.read_file()
        sheet = self.get_users_sheet(excel_file)

        self.validate_columns_order(sheet)

        users = []

        for row in range(2, sheet.max_row + 1):

            user = {
                "email": sheet.cell(row=row, column=1).value,
                "age": sheet.cell(row=row, column=2).value,
                "password": sheet.cell(row=row, column=3).value,
                "score": sheet.cell(row=row, column=4).value,
                "games_played": sheet.cell(row=row, column=5).value,
            }

            users.append(user)

        return users

    def get_users_sheet(self, sheet):
        return sheet["Users"]

    def read_file(self) -> openpyxl.Workbook:
        return openpyxl.load_workbook(self.file_path)

    def validate_columns_order(self, sheet) -> None:
        for column in range(1, sheet.max_column):
            header = sheet[get_column_letter(column) + "1"].value
            if header not in self.headers:
                raise Exception(f"Column {header} should not exist on excel file.")

            if header != self.headers[column - 1]:
                raise Exception(
                    f"Column {header} is not on the right order. The columns order should be {self.headers}"
                )

    def get_user(self, email: str) -> Union[Dict, None]:
        users = self.get_all_users()

        for user in users:
            if user["email"] == email:
                return user

    def save_user(self, user: Dict) -> bool:
        excel_file = self.read_file()
        sheet = self.get_users_sheet(excel_file)

        last_row = sheet.max_row + 1

        sheet.cell(row=last_row, column=1).value = user["email"]
        sheet.cell(row=last_row, column=2).value = user["age"]
        sheet.cell(row=last_row, column=3).value = user["password"]
        sheet.cell(row=last_row, column=4).value = user["score"]
        sheet.cell(row=last_row, column=5).value = user["games_played"]

        try:
            excel_file.save(self.file_path)
        except PermissionError:
            raise PermissionError("Close your excel file so the user can be saved.")

    def update_user_games_played(self, player: Dict) -> bool:
        excel_file = self.read_file()
        sheet = self.get_users_sheet(excel_file)

        self.validate_column_games_played(sheet)

        for row in range(2, sheet.max_row + 1):
            email = sheet.cell(row=row, column=1).value

            if email == player["email"]:
                sheet.cell(row=row, column=5).value = (
                    sheet.cell(row=row, column=5).value + 1
                )
                break

        try:
            excel_file.save(self.file_path)
            return True
        except PermissionError:
            raise PermissionError("Close your excel file so we can update it.")

    def update_user_score(self, player: Dict):
        excel_file = self.read_file()
        sheet = self.get_users_sheet(excel_file)
        self.validate_column_score(sheet)

        for row in range(2, sheet.max_row + 1):
            email = sheet.cell(row=row, column=1).value
            if email == player["email"]:
                sheet.cell(row=row, column=4).value = (
                    sheet.cell(row=row, column=4).value + 1
                )
                break

        try:
            excel_file.save(self.file_path)
            return True
        except PermissionError:
            raise PermissionError("Close your excel file so we can update it.")

    def validate_column_games_played(self, sheet):
        if sheet.cell(row=1, column=5).value != "games played":
            raise Exception(
                f"Column 'games played' is not on the right order. The columns order should be {self.headers}"
            )

    def validate_column_score(self, sheet):
        if sheet.cell(row=1, column=4).value != "score":
            raise Exception(
                f"Column 'score' is not on the right order. The columns order should be {self.headers}"
            )
