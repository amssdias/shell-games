from pathlib import Path
from unittest.mock import MagicMock, Mock, patch

import openpyxl

from models.files_models.excel_file import ExcelFile
from tests.models.utils.excel_file import TestExcelFileUtils


class TestExcelFile(TestExcelFileUtils):
    def setUp(self) -> None:
        super().setUp()
        self.file_directory = Path(Path.cwd(), "tests", "models", "db", "testing.xlsx")
        self.db = ExcelFile(file_path=self.file_directory)

    def test_initial_data(self):
        excel_initial_variables = self.db.__dict__.keys()

        self.assertIn("file_path", excel_initial_variables)
        self.assertIn("headers", excel_initial_variables)
        self.assertIsInstance(self.db.file_path, Path)
        self.assertEqual(self.db.file_path, self.file_directory)

    def test_get_all_users(self):
        self.write_data_to_excel()
        users = self.db.get_all_users()
        self.delete_data_from_excel()

        self.assertCountEqual(users, [self.user_1])

    def test_get_users_sheet(self):
        excel_file = openpyxl.load_workbook(self.file_directory)
        sheet = self.db.get_users_sheet(excel_file)

        self.assertIsInstance(sheet, openpyxl.worksheet.worksheet.Worksheet)

    def test_read_file(self):
        excel_file = self.db.read_file()

        self.assertIsInstance(excel_file, openpyxl.Workbook)

    def test_validate_columns_order(self):
        users_sheet = openpyxl.load_workbook(self.file_directory)["Users"]
        self.assertTrue(self.db.validate_columns_order(users_sheet))

    def _test_validate_columns_order_missing_header(self):
        pass

    def _test_validate_columns_order_wrong_order(self):
        pass

    def test_get_user(self):
        self.write_data_to_excel()
        user = self.db.get_user(self.user_1["email"])
        self.delete_data_from_excel()

        self.assertEqual(user, self.user_1)
        self.assertIsInstance(user, dict)

    def test_get_user_dont_exist(self):
        user = self.db.get_user("testing@bogusemail.com")

        self.assertFalse(user)

    def test_save_user(self):
        user_saved = self.db.save_user(self.user_1)
        self.assertTrue(user_saved)

        sheet = openpyxl.load_workbook(self.file_directory)["Users"]

        confirmed_user = None
        for row in range(2, sheet.max_row + 1):
            confirmed_user = sheet.cell(row=row, column=1).value
            if confirmed_user == self.user_1["email"]:
                break

        self.assertEqual(confirmed_user, self.user_1["email"])

        self.delete_data_from_excel()

    def test_update_user_games_played(self):
        self.write_data_to_excel()
        user = {"email": self.user_1["email"]}

        user_updated = self.db.update_user_games_played(user)
        self.assertTrue(user_updated)

        check_user = self.db.get_user(self.user_1["email"])
        self.assertEqual(check_user["games_played"], self.user_1["games_played"] + 1)

        self.delete_data_from_excel()

    def test_update_user_score(self):
        self.write_data_to_excel()
        user = {"email": self.user_1["email"]}

        user_updated = self.db.update_user_score(user)
        self.assertTrue(user_updated)

        check_user = self.db.get_user(self.user_1["email"])
        self.assertEqual(check_user["score"], self.user_1["score"] + 1)

        self.delete_data_from_excel()

    def test_validate_column_games_played(self):
        sheet = MagicMock()
        cell = MagicMock(value="games played")
        sheet.cell.return_value = cell

        self.db.validate_column_games_played(sheet)

    def test_validate_column_games_played_wrong_column(self):
        sheet = MagicMock()
        cell = MagicMock(value="score")
        sheet.cell.return_value = cell

        with self.assertRaises(Exception) as e:
            self.db.validate_column_games_played(sheet)

    def test_validate_column_games_played_empty_column(self):
        sheet = MagicMock()
        cell = MagicMock(value="")
        sheet.cell.return_value = cell

        with self.assertRaises(Exception) as e:
            self.db.validate_column_games_played(sheet)

    def test_validate_column_score(self):
        pass
