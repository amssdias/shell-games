import openpyxl

from tests.models.utils.file_utils import TestFileUtils


class TestExcelFileUtils(TestFileUtils):
    def write_data_to_excel(self):
        excel_file = openpyxl.load_workbook(self.file_directory)
        sheet = excel_file["Users"]

        last_row = sheet.max_row + 1

        sheet.cell(row=last_row, column=1).value = self.user_1["email"]
        sheet.cell(row=last_row, column=2).value = self.user_1["age"]
        sheet.cell(row=last_row, column=3).value = self.user_1["password"]
        sheet.cell(row=last_row, column=4).value = self.user_1["score"]
        sheet.cell(row=last_row, column=5).value = self.user_1["games_played"]

        excel_file.save(self.file_directory)

    def delete_data_from_excel(self):
        excel_file = openpyxl.load_workbook(self.file_directory)
        sheet = excel_file["Users"]
        sheet.delete_rows(2)
        excel_file.save(self.file_directory)
